#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     :2022-01-13 22:34
# @Author   :wangqinghua
# @Email    : 867075698@qq.com
# @File     : do_excel.py
# @Software : PyCharm


from openpyxl import load_workbook
import pprint
#绝对路径 相对路径

class Do_Excel():
    def get_data(self,file_name,shell_name):
        wb=load_workbook(file_name)
        sheet=wb[shell_name]
        test_data=[] #不断变化的是列,case_id=1,module=2,看excel依次类写
        for i in range(2,sheet.max_row+1):
                row_data={}
                row_data["case_id"]=sheet.cell(i,1).value
                row_data["module"] = sheet.cell(i, 2).value
                row_data["title"] = sheet.cell(i, 3).value
                row_data["url"]=sheet.cell(i,4).value
                row_data["data"] = sheet.cell(i, 5).value
                row_data["expected"] = sheet.cell(i, 6).value
                row_data["method"] = sheet.cell(i, 7).value
                test_data.append(row_data)
        return test_data
    def write_back(self,file_name,sheet_name,i,value):
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(i,8).value=value
        wb.save(file_name)  #保存结果

if __name__ == '__main__':
    res=Do_Excel().get_data("tests.xlsx","test")
    print(res)