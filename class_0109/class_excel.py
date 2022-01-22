#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     :2022-01-09 20:01
# @Author   :wangqinghua
# @Email    : 867075698@qq.com
# @File     : class_excel.py
# @Software : PyCharm


#存到excel里面 python去操作excel

from openpyxl import load_workbook

# 打开工作簿
wb =load_workbook("test.xlsx")

# 定位表单/工作表
sheet = wb["test"]

# 定位单元格
res = sheet.cell(1,1).value
print("拿到的结果是:{}".format(res))
# print("最大行：{}".format(sheet.max_row))
# print("最大列：{}".format(sheet.max_column))

#从excel获取到的数据是什么类型
# 从excel读取数据：数字还是数字 其他都是字符串类型
# print("method:{}，类型是{}".format(sheet.cell(1,1).value,type(sheet.cell(1,1).value)))
# print("url:{}，类型是{}".format(sheet.cell(1,2).value,type(sheet.cell(1,2).value)))
# print("data:{}，类型是{}".format(sheet.cell(1,3).value,type(sheet.cell(1,3).value)))
# print("code:{}，类型是{}".format(sheet.cell(1,4).value,type(sheet.cell(1,4).value)))


# eval() 把数据类 转换成原本的类型 强转
# 把字符串里面的字典 强制转换成字典类型
# s = '{"age":13}'
# print(type(eval(s)))