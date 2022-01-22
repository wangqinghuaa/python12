#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     :2022-01-09 21:06
# @Author   :wangqinghua
# @Email    : 867075698@qq.com
# @File     : do_excel.py
# @Software : PyCharm

from openpyxl import load_workbook

# 打开工作簿
wb =load_workbook("test.xlsx")

# 定位表单/工作表
sheet = wb["test"]

#怎么拿到第一行所有的数据

class DoExcel:
    def __init__(self,file_name,shell_name):
        self.file_name=file_name
        self.shell_name=shell_name
    def get_data(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.shell_name]
        test_data=[]
        for i in range(1,sheet.max_row+1):
            sub_data={}
            sub_data['method']=sheet.cell(i,1).value
            sub_data['url']=sheet.cell(i,2).value
            sub_data['data']=sheet.cell(i,3).value
            sub_data['code']=sheet.cell(i,4).value
            test_data.append(sub_data)
        print(test_data)

if __name__ == '__main__':
    DoExcel("test.xlsx","test").get_data()
