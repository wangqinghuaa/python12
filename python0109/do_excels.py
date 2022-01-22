#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     :2022-01-09 22:58
# @Author   :wangqinghua
# @Email    : 867075698@qq.com
# @File     : do_excels.py
# @Software : PyCharm

from openpyxl import load_workbook


class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name =file_name
        self.sheet_name=sheet_name

    def get_header(self): #获取标题行
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header =[] #储存标题行
        for i in range(1,sheet.max_column+1):
            header.append(sheet.cell(1,i).value)
        return header

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = self.get_header()
        test_data=[]

        for i in range(2,sheet.max_row+1):
            sub_data={}
            for j in range(4,sheet.max_column+1):
                sub_data[header[j-1]] = sheet.cell(i,j).value
            test_data.append(sub_data)
        return test_data

if __name__ == '__main__':
    print(DoExcel("tests.xlsx","test").get_data())