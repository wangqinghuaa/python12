#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     :2022-01-09 20:41
# @Author   :wangqinghua
# @Email    : 867075698@qq.com
# @File     : do_excel.py
# @Software : PyCharm

from openpyxl import load_workbook
from python0109.read_config import Read_Config

class DoExcel:      #带title的
    def __init__(self,file_name,sheet_name):
        self.file_name =file_name
        self.sheet_name=sheet_name
    def get_data(self,mode="all"):

        #从配置文件读取
        mode=Read_Config().read_config("case.config","MODE","mode")
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data=[]
        for i in range(2,sheet.max_row+1):
            sub_data={}
            sub_data["case_id"] = sheet.cell(i, 1).value
            sub_data["module"] = sheet.cell(i, 2).value
            sub_data["title"] = sheet.cell(i, 3).value
            sub_data["url"] = sheet.cell(i, 4).value
            sub_data["data"] = sheet.cell(i, 5).value
            sub_data["expected"] = sheet.cell(i, 6).value
            sub_data["method"]=sheet.cell(i,7).value
            test_data.append(sub_data)


            # 根据 mode=="all"值取进行判断
        if mode=="all":
            final_data=test_data
        else:#[1,2,3,4,.....]
            final_data=[]
            for item in test_data:#对所有的测试数据进行遍历
                if item["case_id"] in eval(mode):
                    final_data.append(item)
        return final_data  #返回获取到的数据


if __name__ == '__main__':
    print(DoExcel("tests.xlsx","test").get_data())